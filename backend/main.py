"""
ICBF Informes - Backend FastAPI
Modos:
  - basico:    General y/o Gestante de forma independiente (al menos uno)
  - completo:  General + Gestante + BeneficiariosPIActivos (análisis de déficit)
"""

import base64
import io
import math
import warnings
from datetime import datetime

import pandas as pd
from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

app = FastAPI(title="ICBF Informes")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# Categorías en orden de presentación, usadas para contar y ordenar la hoja
CATEGORIAS = {
    "general": [
        "Obesidad",
        "Sobrepeso",
        "Riesgo de Sobrepeso",
        "Peso Adecuado para la Talla",
        "Riesgo de Desnutrición Aguda",
        "Desnutrición Aguda Moderada",
        "Desnutrición Aguda Severa",
    ],
    "gestante": [
        "Bajo Peso para la Edad Gestacional",
        "IMC Adecuado para la Edad Gestacional",
        "Sobrepeso para la Edad Gestacional",
        "Obesidad para la Edad Gestacional",
    ],
}

PERFIL = {
    "general": {
        "col_doc":       "Numero Documento Beneficiario",
        "col_diag":      "ESTADO PESO TALLA",
        "hoja_alerta":   "Alerta Desnutricion",
        "tipos_activos": ["MENOR DE SEIS MESES", "NIÑO O NIÑA ENTRE 6 MESES Y 5 AÑOS Y 11 MESES"],
        "col_fecha":     "FECHA VALORACION NURICIONAL",
        "hoja_excel":    "ICBFCUEGeneralPorToma",
    },
    "gestante": {
        "col_doc":       "Número documento beneficiario",
        "col_diag":      "EST.NUTR. GESTANTE",
        "hoja_alerta":   "Alerta Nutricional",
        "tipos_activos": ["PERSONA GESTANTE"],
        "col_fecha":     "FECHA VALORACION NUTRICIONAL",
        "hoja_excel":    "GestanteLactantePorToma",
    },
}

REEMPLAZOS = str.maketrans("ÓÍÁÉÚ\n", "OIAEU ")
MAX_UPLOAD_MB = 30


def norm(col: str) -> str:
    return col.strip().upper().translate(REEMPLAZOS).replace("  ", " ").strip()


def leer_bytes(data: bytes, filename: str, preferir_hoja: str = None) -> pd.DataFrame:
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext in ("xlsx", "xls"):
        xf = pd.ExcelFile(io.BytesIO(data))
        hoja = xf.sheet_names[0]
        if preferir_hoja:
            for h in xf.sheet_names:
                if preferir_hoja.lower() in h.lower():
                    hoja = h
                    break
        return pd.read_excel(io.BytesIO(data), sheet_name=hoja, dtype=str)
    texto = data.decode("utf-8", errors="ignore")
    sep = ";" if texto.split("\n")[0].count(";") > texto.split("\n")[0].count(",") else ","
    return pd.read_csv(io.StringIO(texto), sep=sep, on_bad_lines="skip", low_memory=False, dtype=str)


def tomas_esperadas(meses: float, intervalo: int) -> int:
    return max(1, math.floor(meses / intervalo))


def procesar(
    nut_bytes: bytes, nut_fn: str,
    act_bytes: bytes | None, act_fn: str | None,
    tipo: str, intervalo: int,
) -> list[dict]:
    p    = PERFIL[tipo]
    cats = CATEGORIAS[tipo]

    df = leer_bytes(nut_bytes, nut_fn, preferir_hoja=p["hoja_excel"])
    df.columns = [norm(c) for c in df.columns]

    COL_DOC      = norm(p["col_doc"])
    COL_DIAG     = p["col_diag"]
    COL_FECHA    = p["col_fecha"]
    COL_ESTADO   = "ESTADO"
    COL_CONTRATO = "NUMERO CONTRATO"
    COL_UNIDAD   = "NOMBRE UNIDAD"
    COL_NOMBRE   = "PRIMER NOMBRE BENEFICIARIO"
    COL_APELLIDO = "PRIMER APELLIDO BENEFICIARIO"

    for col in [COL_FECHA, COL_ESTADO, COL_CONTRATO, COL_DOC]:
        if col not in df.columns:
            raise ValueError(f"Columna no encontrada en archivo {tipo}: '{col}'. ¿Es el archivo correcto?")

    df[COL_ESTADO] = df[COL_ESTADO].str.strip().str.upper()
    df = df[df[COL_ESTADO] == "VINCULADO"].copy()

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        df[COL_FECHA] = pd.to_datetime(df[COL_FECHA], errors="coerce", format="mixed", dayfirst=True)

    COL_NTOMA = next(
        (c for c in df.columns if "TOMA DEL BENEFICIARIO" in c or "NUMERO DE TOMA" in c), None
    )
    df_ord = df.dropna(subset=[COL_FECHA]).copy()
    if COL_NTOMA:
        df_ord["_n"] = pd.to_numeric(df_ord[COL_NTOMA], errors="coerce").fillna(0)
        df_ord = df_ord.sort_values([COL_DOC, COL_FECHA, "_n"])
    else:
        df_ord = df_ord.sort_values([COL_DOC, COL_FECHA])
    df_ultimo = df_ord.drop_duplicates(subset=[COL_DOC], keep="last")

    ahora = pd.to_datetime(datetime.now())
    resultados = []

    for contrato in df[COL_CONTRATO].dropna().unique():
        df_c   = df[df[COL_CONTRATO] == contrato]
        df_ult = df_ultimo[df_ultimo[COL_CONTRATO] == contrato]

        # ── Hoja 1: Usuarios por unidad ──────────────────────────────
        hoja_unidades = (
            df_c.groupby(COL_UNIDAD)[COL_DOC]
            .nunique().reset_index()
            .rename(columns={COL_DOC: "TOTAL USUARIOS UNICOS"})
        )

        # ── Hoja 2: Tomas faltantes ───────────────────────────────────
        df_sin_toma = None
        if act_bytes:
            df_act = leer_bytes(act_bytes, act_fn)
            tipos_u = [t.upper() for t in p["tipos_activos"]]
            df_act = df_act[df_act["Nombre Tipo de beneficiario"].str.strip().str.upper().isin(tipos_u)].copy()
            df_act = df_act[df_act["Número del Contrato"].str.strip() == str(contrato).strip()].copy()
            df_act["_doc"]       = df_act["Documento del beneficiario"].str.strip()
            df_act["_vinc"]      = pd.to_datetime(df_act["Fecha de vinculación del beneficiario"], errors="coerce", dayfirst=True)
            df_act["_meses"]     = (ahora.year - df_act["_vinc"].dt.year) * 12 + (ahora.month - df_act["_vinc"].dt.month)
            df_act["_esperadas"] = df_act["_meses"].apply(lambda m: tomas_esperadas(m, intervalo) if pd.notna(m) else 0)
            conteo = df_c.groupby(COL_DOC).size().reset_index(name="_reales")
            df_act = df_act.merge(conteo, left_on="_doc", right_on=COL_DOC, how="left")
            df_act["_reales"] = df_act["_reales"].fillna(0).astype(int)
            faltantes_src = df_act[df_act["_reales"] < df_act["_esperadas"]]

            filas = []
            for _, row in faltantes_src.iterrows():
                doc = row["_doc"]
                reg = df_ult[df_ult[COL_DOC] == doc]
                fecha_val = reg[COL_FECHA].values[0] if not reg.empty and COL_FECHA in reg.columns else None
                fecha_str = pd.Timestamp(fecha_val).strftime("%Y-%m-%d") if fecha_val is not None and pd.notna(fecha_val) else "SIN TOMA"
                filas.append({
                    "UNIDAD":            reg[COL_UNIDAD].values[0] if not reg.empty else row.get("Nombre de la unidad de servicio", ""),
                    "DOCUMENTO":         doc,
                    "NOMBRE":            reg[COL_NOMBRE].values[0] if not reg.empty and COL_NOMBRE in reg.columns else row.get("Primer Nombre del beneficiario", ""),
                    "APELLIDO":          reg[COL_APELLIDO].values[0] if not reg.empty and COL_APELLIDO in reg.columns else row.get("Primer apellido del beneficiario", ""),
                    "TOMAS REALIZADAS":  int(row["_reales"]),
                    "TOMAS ESPERADAS":   int(row["_esperadas"]),
                    "MESES VINCULADO":   int(row["_meses"]) if pd.notna(row["_meses"]) else 0,
                    "FECHA ULTIMA TOMA": fecha_str,
                    "MOTIVO":            "Sin toma registrada" if row["_reales"] == 0 else "Tomas insuficientes",
                })
            df_faltantes = pd.DataFrame(filas)

            filas_st = []
            for _, row in df_act[df_act["_reales"] == 0].iterrows():
                filas_st.append({
                    "UNIDAD":    row.get("Nombre de la unidad de servicio", ""),
                    "DOCUMENTO": row["_doc"],
                    "NOMBRE":    row.get("Primer Nombre del beneficiario", ""),
                    "APELLIDO":  row.get("Primer apellido del beneficiario", ""),
                })
            df_sin_toma = pd.DataFrame(filas_st)
        else:
            ids_faltantes = set()
            for usuario, grupo in df_ord[df_ord[COL_CONTRATO] == contrato].groupby(COL_DOC):
                fechas = grupo[COL_FECHA].dt.to_period("M").unique()
                if len(fechas) > 1:
                    for i in range(1, len(fechas)):
                        diff = (fechas[i].year - fechas[i-1].year) * 12 + (fechas[i].month - fechas[i-1].month)
                        if diff > intervalo:
                            ids_faltantes.add(usuario)
                            break
                if usuario not in ids_faltantes:
                    ultima = fechas[-1]
                    diff_hoy = (ahora.year - ultima.year) * 12 + (ahora.month - ultima.month)
                    if diff_hoy >= intervalo:
                        ids_faltantes.add(usuario)

            cols_f = [c for c in [COL_UNIDAD, COL_DOC, COL_NOMBRE, COL_APELLIDO, COL_FECHA] if c in df_ult.columns]
            df_faltantes = df_ult[df_ult[COL_DOC].isin(ids_faltantes)][cols_f].copy() if ids_faltantes else pd.DataFrame(columns=cols_f)
            if not df_faltantes.empty and COL_FECHA in df_faltantes.columns:
                df_faltantes[COL_FECHA] = df_faltantes[COL_FECHA].dt.strftime("%Y-%m-%d")

        # ── Hoja 3: Estado nutricional (todos los últimos registros) ──
        if COL_DIAG in df_ult.columns:
            cols_a = [c for c in [COL_UNIDAD, COL_DOC, COL_NOMBRE, COL_APELLIDO, COL_DIAG, COL_FECHA] if c in df_ult.columns]
            df_alerta = df_ult[df_ult[COL_DIAG].notna()][cols_a].copy()
            if COL_FECHA in df_alerta.columns:
                df_alerta[COL_FECHA] = df_alerta[COL_FECHA].dt.strftime("%Y-%m-%d")

            # Ordenar por el orden de categorías definido
            cat_order = {norm(c): i for i, c in enumerate(cats)}
            df_alerta["_ord"] = df_alerta[COL_DIAG].apply(lambda x: cat_order.get(norm(str(x)), len(cats)))
            df_alerta = df_alerta.sort_values("_ord").drop(columns=["_ord"])

            # Contar por categoría (comparación normalizada)
            diag_norm = df_ult[COL_DIAG].apply(lambda x: norm(str(x)) if pd.notna(x) else "")
            alertas_conteo = {cat: int((diag_norm == norm(cat)).sum()) for cat in cats}
            # Capturar valores que no corresponden a ninguna categoría conocida
            otros = int(df_ult[COL_DIAG].notna().sum()) - sum(alertas_conteo.values())
            if otros > 0:
                alertas_conteo["Otro"] = otros
        else:
            df_alerta = pd.DataFrame()
            alertas_conteo = {cat: 0 for cat in cats}

        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            hoja_unidades.to_excel(writer, sheet_name="Usuarios por Unidad", index=False)
            (df_faltantes if not df_faltantes.empty
             else pd.DataFrame({"MENSAJE": ["Sin tomas faltantes registradas"]})).to_excel(
                writer, sheet_name="Tomas Faltantes", index=False)
            (df_alerta if not df_alerta.empty
             else pd.DataFrame({"MENSAJE": ["Sin registros nutricionales"]})).to_excel(
                writer, sheet_name=p["hoja_alerta"], index=False)
            if df_sin_toma is not None:
                (df_sin_toma if not df_sin_toma.empty
                 else pd.DataFrame({"MENSAJE": ["Todos los beneficiarios tienen tomas registradas"]})).to_excel(
                    writer, sheet_name="Sin Tomas Registradas", index=False)
        buf.seek(0)

        resultados.append({
            "contrato":  str(contrato),
            "tipo":      tipo,
            "vinculados": int(df_c[COL_DOC].nunique()),
            "faltantes":  len(df_faltantes),
            "alertas":    alertas_conteo,
            "unidades":   len(hoja_unidades),
            "sin_toma":   len(df_sin_toma) if df_sin_toma is not None else 0,
            "filename":   f"Informe_{tipo.upper()}_Contrato_{contrato}.xlsx",
            "bytes":      buf.read(),
        })

    return resultados


# ─────────────────────────────────────────────
# REPORTE BENEFICIARIOS
# ─────────────────────────────────────────────

def generar_reporte(data_bytes: bytes, filename: str) -> tuple[bytes, list[dict], dict]:
    df = leer_bytes(data_bytes, filename, preferir_hoja="ICBFCUEBeneficiariosPIActivosRe")

    def find_col(*keywords):
        for kw in keywords:
            for c in df.columns:
                if kw.lower() in c.lower():
                    return c
        return None

    col_uds  = find_col("nombre de la unidad de servicio")
    col_tipo = find_col("nombre tipo de beneficiario")
    col_sexo = find_col("sexo del beneficiario")
    col_edad = find_col("edad del beneficiario")

    for label, col in [("Unidad de servicio", col_uds), ("Tipo beneficiario", col_tipo),
                        ("Sexo", col_sexo), ("Edad", col_edad)]:
        if col is None:
            raise ValueError(f"Columna no encontrada: '{label}'")

    df[col_edad] = pd.to_numeric(df[col_edad], errors="coerce").fillna(0)
    df[col_tipo] = df[col_tipo].str.strip().str.upper()
    df[col_sexo] = df[col_sexo].str.strip().str.upper()

    TIPO_MENOR = "MENOR DE SEIS MESES"
    TIPO_NINO  = "NIÑO O NIÑA ENTRE 6 MESES Y 5 AÑOS Y 11 MESES"
    TIPO_GEST  = "PERSONA GESTANTE"

    udses = sorted(df[col_uds].dropna().unique())
    filas = []

    for uds in udses:
        dfu = df[df[col_uds] == uds]

        menor = dfu[dfu[col_tipo] == TIPO_MENOR]
        m_h   = int((menor[col_sexo] == "HOMBRE").sum())
        m_m   = int((menor[col_sexo] == "MUJER").sum())
        m_tot = m_h + m_m

        nino    = dfu[dfu[col_tipo] == TIPO_NINO]
        nh      = nino[nino[col_sexo] == "HOMBRE"]
        nm      = nino[nino[col_sexo] == "MUJER"]
        n_h_lt2 = int((nh[col_edad] < 2).sum())
        n_h_2a5 = int((nh[col_edad] >= 2).sum())
        n_h_tot = n_h_lt2 + n_h_2a5
        n_m_lt2 = int((nm[col_edad] < 2).sum())
        n_m_2a5 = int((nm[col_edad] >= 2).sum())
        n_m_tot = n_m_lt2 + n_m_2a5
        n_tot   = n_h_tot + n_m_tot

        gest  = dfu[dfu[col_tipo] == TIPO_GEST]
        g_tot = len(gest)

        total = m_tot + n_tot + g_tot

        filas.append({
            "uds": uds,
            "m_h": m_h, "m_m": m_m, "m_tot": m_tot,
            "n_h_lt2": n_h_lt2, "n_h_2a5": n_h_2a5, "n_h_tot": n_h_tot,
            "n_m_lt2": n_m_lt2, "n_m_2a5": n_m_2a5, "n_m_tot": n_m_tot,
            "n_tot": n_tot,
            "g_tot": g_tot,
            "total": total,
        })

    keys = [k for k in filas[0] if k != "uds"]
    totales = {k: sum(f[k] for f in filas) for k in keys}
    totales["uds"] = "Total general"

    # ── Construir Excel ────────────────────────────────────────────
    wb  = Workbook()
    ws  = wb.active
    ws.title = "Reporte"

    C_GREEN1 = "0A5C36"
    C_GREEN2 = "1A6B3A"
    C_GREEN3 = "2E7D52"
    C_LIGHT  = "E8F5EE"
    C_HEADER = "F0F2EE"

    def hfill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def hfont(bold=True, color="FFFFFF", size=9):
        return Font(bold=bold, color=color, size=size)

    thin   = Side(style="thin", color="D6D3CC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center")

    def sc(cell_ref, value="", fill=None, font=None, align=None, brd=True):
        c = ws[cell_ref]
        c.value = value
        if fill:  c.fill = fill
        if font:  c.font = font
        if align: c.alignment = align
        if brd:   c.border = border
        return c

    # ── Fila 1: título ──
    ws.merge_cells("A1:K1")
    sc("A1", "Beneficiarios PI Activos por Unidad de Servicio",
       font=Font(bold=True, size=11, color="1A1916"),
       fill=hfill(C_HEADER), align=center, brd=False)
    ws.row_dimensions[1].height = 22

    # ── Fila 2: grupos de tipo ──
    ws.merge_cells("A2:A4")
    sc("A2", "Unidad de Servicio",
       fill=hfill(C_HEADER), font=hfont(color="1A1916"), align=center)

    ws.merge_cells("B2:D2")
    sc("B2", "MENOR DE SEIS MESES", fill=hfill(C_GREEN1), font=hfont(), align=center)

    ws.merge_cells("E2:I2")
    sc("E2", "NIÑO O NIÑA ENTRE 6 MESES Y 5 AÑOS Y 11 MESES",
       fill=hfill(C_GREEN2), font=hfont(), align=center)

    ws.merge_cells("J2:J4")
    sc("J2", "PERSONA GESTANTE", fill=hfill(C_GREEN3), font=hfont(), align=center)

    ws.merge_cells("K2:K4")
    sc("K2", "Total general", fill=hfill(C_GREEN1), font=hfont(), align=center)

    ws.row_dimensions[2].height = 32

    # ── Fila 3: sexo ──
    # MENOR: H, M, Total (sin sub-rangos de edad — todos son < 6 meses)
    ws.merge_cells("B3:B4")
    sc("B3", "Hombre", fill=hfill(C_GREEN1), font=hfont(), align=center)
    ws.merge_cells("C3:C4")
    sc("C3", "Mujer", fill=hfill(C_GREEN1), font=hfont(), align=center)
    ws.merge_cells("D3:D4")
    sc("D3", "Total", fill=hfill(C_GREEN1), font=hfont(), align=center)

    # NIÑO: Hombre (2 sub-rangos) | Mujer (2 sub-rangos) | Total
    ws.merge_cells("E3:F3")
    sc("E3", "Hombre", fill=hfill(C_GREEN2), font=hfont(), align=center)
    ws.merge_cells("G3:H3")
    sc("G3", "Mujer", fill=hfill(C_GREEN2), font=hfont(), align=center)
    ws.merge_cells("I3:I4")
    sc("I3", "Total", fill=hfill(C_GREEN2), font=hfont(), align=center)

    ws.row_dimensions[3].height = 18

    # ── Fila 4: sub-rangos de edad solo para NIÑO ──
    for ref, lbl in [("E4", "<2 años"), ("F4", "2-5 años"), ("G4", "<2 años"), ("H4", "2-5 años")]:
        sc(ref, lbl, fill=hfill(C_GREEN2), font=hfont(size=8), align=center)

    ws.row_dimensions[4].height = 16

    # ── Filas de datos ──
    all_rows = filas + [totales]
    for i, fila in enumerate(all_rows):
        r    = i + 5
        is_t = i == len(filas)
        row_fill = hfill(C_LIGHT) if is_t else None
        row_font = hfont(color="1A1916") if is_t else Font(size=9)

        def v(n):
            return n if n != 0 else None

        row_vals = [
            fila["uds"],
            v(fila["m_h"]), v(fila["m_m"]), v(fila["m_tot"]),
            v(fila["n_h_lt2"]), v(fila["n_h_2a5"]),
            v(fila["n_m_lt2"]), v(fila["n_m_2a5"]),
            v(fila["n_tot"]),
            v(fila["g_tot"]),
            v(fila["total"]),
        ]

        for j, val in enumerate(row_vals):
            col_ref = f"{get_column_letter(j + 1)}{r}"
            c = ws[col_ref]
            c.value = val
            c.font  = row_font
            c.alignment = left if j == 0 else center
            c.border = border
            if row_fill:
                c.fill = row_fill

        ws.row_dimensions[r].height = 15

    # ── Anchos de columna ──
    ws.column_dimensions["A"].width = 32
    for col_l in ["B","C","D","E","F","G","H","I","J","K"]:
        ws.column_dimensions[col_l].width = 9

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read(), filas, totales


# ─────────────────────────────────────────────
# ENDPOINTS
# ─────────────────────────────────────────────

@app.get("/health")
def health():
    return {"status": "ok", "timestamp": datetime.now().isoformat()}


@app.post("/reporte-beneficiarios")
async def reporte_beneficiarios(activos: UploadFile = File(...)):
    data = await activos.read()
    if len(data) > MAX_UPLOAD_MB * 1024 * 1024:
        raise HTTPException(413, f"El archivo supera los {MAX_UPLOAD_MB} MB.")
    try:
        excel_bytes, filas, totales = generar_reporte(data, activos.filename)
    except ValueError as e:
        raise HTTPException(422, str(e))
    except Exception as e:
        raise HTTPException(500, f"Error procesando el archivo: {str(e)}")
    return JSONResponse({
        "data":     base64.b64encode(excel_bytes).decode("ascii"),
        "filename": "Reporte_Beneficiarios_Activos.xlsx",
        "filas":    filas,
        "totales":  totales,
    })


@app.post("/procesar")
async def procesar_archivos(
    modo: str = Form(...),
    general: UploadFile = File(None),
    gestante: UploadFile = File(None),
    activos: UploadFile = File(None),
    intervalo: int = Form(3),
):
    if modo not in ("basico", "completo"):
        raise HTTPException(400, "modo debe ser 'basico' o 'completo'")

    def tiene(f: UploadFile) -> bool:
        return f is not None and bool(f.filename)

    general_bytes  = (await general.read())  if tiene(general)  else None
    gestante_bytes = (await gestante.read()) if tiene(gestante) else None
    act_bytes      = (await activos.read())  if tiene(activos)  else None
    act_fn         = activos.filename        if tiene(activos)  else None

    if modo == "completo":
        if not general_bytes or not gestante_bytes or not act_bytes:
            raise HTTPException(400, "Modo completo requiere los tres archivos.")
    else:
        if not general_bytes and not gestante_bytes:
            raise HTTPException(400, "Debes cargar al menos un archivo (General o Gestante/Lactante).")

    for nombre, data in [("general", general_bytes), ("gestante", gestante_bytes), ("activos", act_bytes)]:
        if data and len(data) > MAX_UPLOAD_MB * 1024 * 1024:
            raise HTTPException(413, f"El archivo {nombre} supera los {MAX_UPLOAD_MB} MB permitidos.")

    try:
        todos = []
        pares = []
        if general_bytes:
            pares.append(("general",  general_bytes,  general.filename))
        if gestante_bytes:
            pares.append(("gestante", gestante_bytes, gestante.filename))
        for tipo, data, fn in pares:
            todos.extend(procesar(data, fn, act_bytes, act_fn, tipo, intervalo))
    except ValueError as e:
        raise HTTPException(422, str(e))
    except Exception as e:
        raise HTTPException(500, f"Error procesando archivos: {str(e)}")

    return JSONResponse({
        "archivos": [
            {
                "filename":  r["filename"],
                "contrato":  r["contrato"],
                "tipo":      r["tipo"],
                "vinculados": r["vinculados"],
                "faltantes":  r["faltantes"],
                "alertas":    r["alertas"],
                "unidades":   r["unidades"],
                "sin_toma":   r["sin_toma"],
                "data":       base64.b64encode(r["bytes"]).decode("ascii"),
            }
            for r in todos
        ]
    })


app.mount("/", StaticFiles(directory="/app/frontend", html=True), name="frontend")
